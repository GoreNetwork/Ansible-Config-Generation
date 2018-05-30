from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

def write_excel_data(row,column,value,sheet):
	tmp = str(get_column_letter(column))+str(row)
	#print (tmp)
	#print (value)
	sheet[tmp] = value
	return column+1

def pull_sheet_names (work_book):
	return (work_book.get_sheet_names())

def pull_cell(sheet, row, cell_letter):
	cell = cell_letter + str(row)
	new_device = sheet[cell].value
	return new_device




def pull_devices_2018(filename, sheet, sitename):
	# print (filename)
	# print (sheet)
	# print (sitename)
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_sheet_by_name(sheet)
	last_cell = 'D' + str(sheet.max_row)
	# print (last_cell)
	rows_i_want = []
	all_devices = []
	for rowOfCellObjects in sheet['D2':last_cell]:
		# print (rowOfCellObjects)
		for cellObj in rowOfCellObjects:
			if cellObj.value == sitename:
				# print ("Row = "+str(cellObj.row)+ " Column = "+(cellObj.column))
				rows_i_want.append(cellObj.row)
	for row in rows_i_want:
		# Pull the IP
		ip = pull_cell(sheet, row, "B")
		# print (ip)
		# Current device type
		old_device = pull_cell(sheet, row, "F")
		# New Device Type
		new_device = pull_cell(sheet, row, "R")
		# Sales Number
		sale_num = str(pull_cell(sheet, row, "R"))
		cell = "BF" + str(row)
		sale_num = sheet[cell].value
		# Site Name
		site_name = pull_cell(sheet, row, "A")
		# Addr
		addr = pull_cell(sheet, row, "R")
		# City
		city = pull_cell(sheet, row, "I")
		state = pull_cell(sheet, row, "H")
		zip = pull_cell(sheet, row, "R")
		full_address = pull_cell(sheet, row, "R")

		device_swap = [ip, old_device, new_device, sale_num, site_name, addr, city, state, zip, full_address]
		all_devices.append(device_swap)

	return (all_devices)
