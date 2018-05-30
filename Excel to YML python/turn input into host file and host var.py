from excel_work import *
import openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from pprint import pprint
from common_functions import *

def make_host_vars_file(dic,rows):
   # pprint (dic)
    output_file = dic['hostname']+'.yml'
    pprint (output_file)
    start_of_host_vars = "---\n"
    final_doc = start_of_host_vars
    switches = []
    routers = []
    for row in rows:
        key = row[0]
        value = dic[key]
        final_doc = final_doc+"{}: {}\n".format (key,value)
    to_doc_w(output_file, final_doc)

def build_hosts_file(dic_list):
    routers = []
    switches = []




input_file = "input.xlsx"
hosts_file = 'hosts'
sheet = 'Sheet1'



wb = openpyxl.load_workbook(input_file)
sheet = wb.get_sheet_by_name(sheet)
#Values that will be in the user_var file, and what line they are on in the excel file
rows = [
    ["switch_or_router", "1"],
    ["region", "2"],
    ["management", "3"],
    ["l2_l3", "4"],
    ["dfgw", "5"],
    ["dfgws_snm", "6"],
    ["routing_protocol", "7"],
    ["l3_eigrp-as_ospf_area", "8"],
    ["nac", "9"],
    ["hardware_type", "10"],
    ["stp_priority", "11"],
    ["uplink1_hostname", "12"],
    ["management_vlan", "13"],
    ["voice_vlan", "14"],
    ["data_vlan", "15"],
    ["guest_vlan", "16"],
    ["hostname", "17"],
    ["cca", "18"],

]

columns_i_want = sheet.max_column
#pprint (columns_i_want)
devices = []
for column in range(2,columns_i_want+1):
    column_letter = get_column_letter(column)
    tmp = {}
    for row_values in rows:
        value,row = row_values
       # print (row,value)
        tmp[value]=pull_cell(sheet, row, column_letter)
    devices.append(tmp)


#pprint (devices)

for device in devices:
    make_host_vars_file(device, rows)
